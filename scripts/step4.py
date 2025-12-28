from openpyxl import load_workbook
from datetime import datetime

FILE = r"C:\Users\DELL\Desktop\GFS_Automation\input_files\Delta.xlsx"

# =====================================================
# STEP 1 — PROCESS 2025 SHEET
# =====================================================

wb = load_workbook(FILE)
ws = wb["2025"]

# -----------------------------
# 1. Read all data once
# -----------------------------
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = rows[1:]

date_col_index = header.index("Date")  # Column A

# -----------------------------
# 2. Filter only 2025 rows
# -----------------------------
filtered = []

for row in data:
    cell = row[date_col_index]

    if cell is None:
        continue

    # Works for both datetime and string
    if isinstance(cell, str):
        try:
            year = int(cell[:4])
        except:
            continue
    else:
        year = cell.year

    if year == 2025:
        filtered.append(row)

# -----------------------------
# 3. Clear sheet completely
# -----------------------------
ws.delete_rows(1, ws.max_row)

# -----------------------------
# 4. Write header + new Aging column
# -----------------------------
new_header = list(header)
new_header.insert(1, "Aging")
ws.append(new_header)

# -----------------------------
# 5. Write data + formula
# -----------------------------
for i, row in enumerate(filtered, start=2):
    row = list(row)
    row.insert(1, f"=EOMONTH(A{i},0)")
    ws.append(row)

# -----------------------------
# 6. Save safely
# -----------------------------
wb.save(FILE)

print("✅ DONE — Clean, fast, zero errors.")




# =====================================================
# STEP 2 — PROCESS GFS SHEET
# =====================================================
ws_gfs = wb["GFS"]

# Insert Aging column
ws_gfs.insert_cols(2)
ws_gfs["B1"] = "Aging"

for row in range(2, ws_gfs.max_row + 1):
    ws_gfs[f"B{row}"] = (
        f'=IF(TODAY()-A{row}<=30,"1-30",'
        f'IF(TODAY()-A{row}<=60,"31-60",'
        f'IF(TODAY()-A{row}<=90,"61-90","90+")))'
    )

# ---------------- SAVE FILE ----------------
wb.save(FILE_PATH)

print("✅ 2025 filtering + Aging columns added successfully.")
