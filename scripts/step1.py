import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ---------------- PATHS ----------------
BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input_files"

CASE_FILE = INPUT_DIR / "CombinedOwnerCase-AR.xlsx"
SUPPORTING_FILE = INPUT_DIR / "Supporting.xlsx"
FORMAT_FILE = INPUT_DIR / "Format.xlsx"

# ---------------- SETTINGS ----------------
SHEET_NAME = 0  # First sheet in CombineOwnerCase-AR.xlsx

USE_COLS = [
    "CASEID", "PATIENTNAME", "MRN", "BUSINESSREGID", "PROVIDER",
    "OWNER NAME", "CASETYPE", "BILLED", "PAID", "WRITEOFF", "ADVANCE", "OUTSTANDING"
]

# ---------------- LOAD SOURCE DATA ----------------
print("Reading CombineOwnerCase-AR.xlsx ...")

df = pd.read_excel(
    CASE_FILE,
    sheet_name=SHEET_NAME,
    usecols=USE_COLS,
    engine="openpyxl"
)

# ---------------- FILTER DATA ----------------
df = df[df["OWNER NAME"].isin(["GFS", "GFS-SPV", "Med Legal"])]
df = df[df["CASETYPE"].isin([
    "Alternative Dispute Resolution (ADR)",
    "Copy Record (ML)",
    "Workers Compensation"
])]

# ---------------- SPLIT DATA ----------------
group_cols = ["CASEID", "PATIENTNAME", "MRN", "BUSINESSREGID", "PROVIDER"]
finance_cols = ["BILLED", "PAID", "WRITEOFF", "ADVANCE", "OUTSTANDING"]

supporting_df = df[group_cols]
format_df = df[group_cols + finance_cols]

# ---------------- FUNCTION TO CLEAR SHEET ----------------
def clear_sheet_except_header(file_path, sheet_name):
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path.name}")

    ws = wb[sheet_name]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    return wb

# ---------------- UPDATE SUPPORTING FILE ----------------
print("Updating Supporting.xlsx...")
wb_support = clear_sheet_except_header(SUPPORTING_FILE, "GFS Cases")

for r_idx, row in enumerate(supporting_df.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        wb_support["GFS Cases"].cell(row=r_idx, column=c_idx, value=value)

wb_support.save(SUPPORTING_FILE)

# ---------------- UPDATE FORMAT FILE ----------------
print("Updating Format.xlsx...")
wb_format = clear_sheet_except_header(FORMAT_FILE, "Sheet1")

for r_idx, row in enumerate(format_df.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        wb_format["Sheet1"].cell(row=r_idx, column=c_idx, value=value)

wb_format.save(FORMAT_FILE)

print("✅ SUCCESS: Files updated without touching other sheets.")
