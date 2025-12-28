import pandas as pd
from openpyxl import load_workbook

FILE_PATH = r"C:\Users\DELL\Desktop\GFS_Automation\input_files\PaymentDetailReport.xlsx"

# =====================================================
# STEP 1 — PROCESS Monthly-Stats SHEET
# =====================================================

# -------------------- LOAD WORKBOOK --------------------
wb = load_workbook(FILE_PATH)

# -------------------- DELETE UNWANTED SHEETS --------------------
for sheet in ["Summary", "Monthly Stats"]:
    if sheet in wb.sheetnames:
        del wb[sheet]

# -------------------- LOAD PAYMENT REPORT --------------------
df = pd.read_excel(FILE_PATH, sheet_name="Payment Report")

# -------------------- REMOVE 'PI' CASES --------------------
# Keep everything EXCEPT PI
df = df[df["CASETYPE"] != "PI"]

# -------------------- WRITE BACK CLEAN DATA --------------------
with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name="Payment Report", index=False)

print("✅ DONE — 'PI' cases removed successfully.")
